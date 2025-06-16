import openpyxl


class HomePageData:
    test_HomePage_data = [{"firstname": "Name", "lastname": "Surename", "gender": "Male"},
                          {"firstname": "Rahul", "lastname": "shetty", "gender": "Female"}]
    @staticmethod
    def getData(test_case_name):

        book = openpyxl.load_workbook("C:\\Users\\dispolatov\\PycharmProjects\\Excel\\PythonDemo.xlsx")
        sheet = book.active
        mydict = {}
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    mydict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [mydict]
