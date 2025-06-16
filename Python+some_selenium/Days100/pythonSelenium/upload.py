import os
import time
from telnetlib import EC
import openpyxl
from selenium import webdriver
from selenium import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support import WebDriverWait

FILE_PATH = "C:\\Users\\dispolatov\\Downloads\\download.xlsx"
FRUIT_NAME = "Apple"
NEW_PRICE = "800"


def edit_excel(fruit, price):
    book = openpyxl.load_workbook(FILE_PATH)
    sheet = book.active
    fruit_column = -1
    pr_column = -1
    for i in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=i).value == "fruit_name":
            fruit_column = i
        if sheet.cell(row=1, column=i).value == "price":
            pr_column = i
    print(fruit_column)
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column=fruit_column).value == fruit:
            sheet.cell(row=i, column=pr_column).value = price

    book.save(FILE_PATH)


driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.ID, "downloadButton").click()
time.sleep(5)

edit_excel(FRUIT_NAME, NEW_PRICE)

file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(FILE_PATH)

toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait = WebDriverWait(driver, 5)
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

price_column = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH, "//div[text()='"+FRUIT_NAME+"']/parent::div/parent::div/div[@id='cell-"+price_column+"-undefined']").text

assert actual_price == NEW_PRICE


os.remove(FILE_PATH)